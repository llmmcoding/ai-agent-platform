"""
File Reader Tool - Read various file formats
Supports: PDF, DOCX, TXT, MD, CSV, JSON, XML, Excel
- Local file and URL support
- File size limit (10MB)
- Metadata extraction
- ReAct-compatible error format
"""
import json
import os
import logging
from typing import Dict, Optional
from pathlib import Path
from enum import Enum

logger = logging.getLogger(__name__)


class FileFormat(Enum):
    """Supported file formats"""
    PDF = "pdf"
    DOCX = "docx"
    TXT = "txt"
    MD = "md"
    CSV = "csv"
    JSON = "json"
    XML = "xml"
    XLSX = "xlsx"
    UNKNOWN = "unknown"


# Supported file size limit: 10MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Allowed base directories for file reading (security)
ALLOWED_BASE_DIRS = [
    os.path.join(os.path.expanduser("~"), "documents"),
    "/tmp",
    "/var/data",
]


def _is_path_safe(file_path: str) -> bool:
    """
    Check if the file path is safe (no path traversal attacks).
    Returns True if the resolved path is within allowed directories.
    """
    try:
        # Resolve the real path (resolves symlinks and ..)
        real_path = os.path.realpath(file_path)

        # Check if path contains any traversal patterns
        if ".." in file_path:
            return False

        # Normalize the path
        normalized = os.path.normpath(file_path)

        # Check for absolute path traversal
        if os.path.isabs(normalized):
            # For absolute paths, check if they're under any allowed base dir
            for base_dir in ALLOWED_BASE_DIRS:
                if os.path.commonpath([real_path, base_dir]) == base_dir:
                    return True
            return False

        # For relative paths, check if resolved path is under allowed base dirs
        for base_dir in ALLOWED_BASE_DIRS:
            if os.path.commonpath([real_path, base_dir]) == base_dir:
                return True

        # If no base dir matches, allow in current working directory
        cwd = os.getcwd()
        return os.path.commonpath([real_path, cwd]) == cwd

    except (ValueError, OSError):
        return False


def file_reader_executor(input_data: dict) -> str:
    """
    File reader executor

    Args:
        input_data: Dict containing:
            - file_path: str (required) - Path to file (local or URL)
            - max_chars: int (optional) - Maximum characters to read
            - extract_metadata: bool (optional) - Include file metadata

    Returns:
        JSON string with file content and metadata
    """
    file_path = input_data.get("file_path", "")
    max_chars = min(input_data.get("max_chars", 50000), 100000)
    extract_metadata = input_data.get("extract_metadata", True)

    if not file_path:
        return json.dumps({
            "status": "error",
            "error_code": "INVALID_PARAMETER",
            "message": "file_path is required"
        }, ensure_ascii=False)

    try:
        # Determine if local file or URL
        if file_path.startswith("http://") or file_path.startswith("https://"):
            return _read_from_url(file_path, max_chars, extract_metadata)
        else:
            return _read_local_file(file_path, max_chars, extract_metadata)

    except FileNotFoundError:
        return json.dumps({
            "status": "error",
            "error_code": "FILE_NOT_FOUND",
            "message": f"File not found: {file_path}"
        }, ensure_ascii=False)
    except PermissionError:
        return json.dumps({
            "status": "error",
            "error_code": "PERMISSION_DENIED",
            "message": f"Permission denied: {file_path}"
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"File read failed: {e}")
        return json.dumps({
            "status": "error",
            "error_code": "READ_ERROR",
            "message": f"Failed to read file: {str(e)}"
        }, ensure_ascii=False)


def _detect_format(filename: str) -> FileFormat:
    """Detect file format from extension"""
    ext = Path(filename).suffix.lower().lstrip(".")

    format_map = {
        "pdf": FileFormat.PDF,
        "docx": FileFormat.DOCX,
        "txt": FileFormat.TXT,
        "md": FileFormat.MD,
        "markdown": FileFormat.MD,
        "csv": FileFormat.CSV,
        "json": FileFormat.JSON,
        "xml": FileFormat.XML,
        "xlsx": FileFormat.XLSX,
        "xls": FileFormat.XLSX,
    }

    return format_map.get(ext, FileFormat.UNKNOWN)


def _read_local_file(file_path: str, max_chars: int, extract_metadata: bool) -> str:
    """Read content from local file"""
    # Security check: prevent path traversal attacks
    if not _is_path_safe(file_path):
        return json.dumps({
            "status": "error",
            "error_code": "INVALID_PATH",
            "message": "Path traversal attempts are not allowed"
        }, ensure_ascii=False)

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(file_path)

    file_size = path.stat().st_size

    if file_size > MAX_FILE_SIZE:
        return json.dumps({
            "status": "error",
            "error_code": "FILE_TOO_LARGE",
            "message": f"File exceeds maximum size of {MAX_FILE_SIZE // (1024 * 1024)}MB"
        }, ensure_ascii=False)

    file_format = _detect_format(file_path)

    # Read based on format
    try:
        if file_format == FileFormat.TXT or file_format == FileFormat.MD:
            content = path.read_text(encoding="utf-8", errors="replace")
        elif file_format == FileFormat.JSON:
            import json as json_lib
            data = json_lib.loads(path.read_text(encoding="utf-8"))
            content = json_lib.dumps(data, indent=2, ensure_ascii=False)
        elif file_format == FileFormat.CSV:
            content = _read_csv(path)
        elif file_format == FileFormat.PDF:
            content = _read_pdf(path)
        elif file_format == FileFormat.DOCX:
            content = _read_docx(path)
        elif file_format == FileFormat.XLSX:
            content = _read_excel(path)
        else:
            # Try to read as text
            try:
                content = path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                content = f"[Binary file: {file_format.value}]"

        # Truncate if needed
        truncated = False
        if len(content) > max_chars:
            content = content[:max_chars]
            truncated = True

        result = {
            "status": "success",
            "filename": path.name,
            "format": file_format.value,
            "content": content,
            "size_bytes": file_size,
            "truncated": truncated
        }

        if extract_metadata:
            result["metadata"] = {
                "modified_time": path.stat().st_mtime,
                "is_readable": os.access(file_path, os.R_OK),
                "encoding": "utf-8"
            }

        return json.dumps(result, ensure_ascii=False)

    except Exception as e:
        raise RuntimeError(f"Failed to read file: {str(e)}")


def _read_csv(path: Path) -> str:
    """Read CSV file as formatted text table"""
    import csv

    lines = []
    with path.open(encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                # Header row
                lines.append(" | ".join(row))
                lines.append(" | ".join(["---"] * len(row)))
            else:
                lines.append(" | ".join(row))

            if i >= 100:  # Limit to 100 rows
                lines.append(f"... ({sum(1 for _ in reader)} more rows)")
                break

    return "\n".join(lines)


def _read_pdf(path: Path) -> str:
    """Read PDF file using pypdf"""
    try:
        from pypdf import PdfReader

        reader = PdfReader(path)
        text_parts = []

        for i, page in enumerate(reader.pages[:20]):  # Max 20 pages
            text_parts.append(f"--- Page {i + 1} ---\n{page.extract_text()}")

        return "\n\n".join(text_parts)
    except ImportError:
        return "[PDF reading requires 'pypdf' library. Install with: pip install pypdf]"
    except Exception as e:
        return f"[PDF reading failed: {str(e)}]"


def _read_docx(path: Path) -> str:
    """Read DOCX file using python-docx"""
    try:
        from docx import Document

        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        return "[DOCX reading requires 'python-docx' library. Install with: pip install python-docx]"
    except Exception as e:
        return f"[DOCX reading failed: {str(e)}]"


def _read_excel(path: Path) -> str:
    """Read Excel file using openpyxl"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames[:5]:  # Max 5 sheets
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True, max_row=100)):
                if all(cell is None for cell in row):
                    continue
                rows.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
                if i >= 99:
                    rows.append(f"... ({ws.max_row - 100} more rows)")
                    break
            sheets.append(f"=== {sheet_name} ===\n" + "\n".join(rows))

        return "\n\n".join(sheets)
    except ImportError:
        return "[Excel reading requires 'openpyxl' library. Install with: pip install openpyxl]"
    except Exception as e:
        return f"[Excel reading failed: {str(e)}]"


def _read_from_url(url: str, max_chars: int, extract_metadata: bool) -> str:
    """Read file from URL"""
    import httpx

    try:
        with httpx.Client(timeout=30.0) as client:
            response = client.get(url)
            response.raise_for_status()

            content = response.text[:max_chars]
            truncated = len(response.text) > max_chars

            result = {
                "status": "success",
                "filename": url.split("/")[-1].split("?")[0],
                "format": _detect_format(url).value,
                "content": content,
                "size_bytes": len(response.content),
                "truncated": truncated,
                "url": url
            }

            if extract_metadata:
                result["metadata"] = {
                    "content_type": response.headers.get("content-type", "unknown"),
                    "response_code": response.status_code
                }

            return json.dumps(result, ensure_ascii=False)

    except httpx.TimeoutException:
        return json.dumps({
            "status": "error",
            "error_code": "TIMEOUT",
            "message": f"Request timed out after 30 seconds"
        }, ensure_ascii=False)
    except httpx.HTTPStatusError as e:
        return json.dumps({
            "status": "error",
            "error_code": "HTTP_ERROR",
            "message": f"HTTP {e.response.status_code}: {str(e)}"
        }, ensure_ascii=False)
    except Exception as e:
        return json.dumps({
            "status": "error",
            "error_code": "DOWNLOAD_ERROR",
            "message": f"Failed to download: {str(e)}"
        }, ensure_ascii=False)
